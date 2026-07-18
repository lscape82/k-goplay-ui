# -*- coding: utf-8 -*-
"""엘리베이터 광고 7개 원본(엑셀)을 네트워크 단위 집계 JSON으로 정규화한다.
원본은 배포 저장소에 포함되지 않으므로(용량·기밀) 로컬 경로에서 읽어 data/elevator-networks.json 생성.
버스 정류장(집계형) 처리와 동일한 철학: 개별 핀이 아니라 네트워크·지역 집계.
"""
import json, os, sys
import openpyxl
import xlrd

SRC = os.environ.get("ELEV_SRC", r"E:\2. 광고플레이(주)\★★★플랫폼개발2_2026_업데이트_언니,부장님\★★★플랫폼UI리모델링\2026 UI활용데이터\엘리베이터")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "elevator-networks.json")

# 시도 정규화 + 대표 좌표(집계 버블 위치)
CENTROIDS = {
    "서울": [37.5665, 126.9780], "경기": [37.4000, 127.2000], "인천": [37.4563, 126.7052],
    "부산": [35.1796, 129.0756], "대구": [35.8714, 128.6014], "광주": [35.1595, 126.8526],
    "대전": [36.3504, 127.3845], "울산": [35.5384, 129.3114], "세종": [36.4801, 127.2890],
    "강원": [37.8228, 128.1555], "충북": [36.6357, 127.4917], "충남": [36.6588, 126.6728],
    "전북": [35.7175, 127.1530], "전남": [34.8161, 126.4630], "경북": [36.5760, 128.5056],
    "경남": [35.4606, 128.2132], "제주": [33.4996, 126.5312],
}
OFFICE_DISTRICTS = {  # 오피스 권역 → 시도 근사
    "CBD": "서울", "GBD": "서울", "YBD": "서울", "YBD/마포": "서울", "BBD": "경기",
    "마포": "서울", "서울": "서울", "수도권": "경기", "동부": "서울", "판교": "경기", "분당": "경기",
}

def norm_sido(v):
    if not v:
        return None
    s = str(v).strip()
    for token in ["특별자치도", "특별자치시", "특별시", "광역시", "특별도"]:
        s = s.replace(token, "")
    s = s.replace("도", "") if s.endswith("도") else s
    alias = {"강원특별": "강원", "전북특별": "전북", "제주특별": "제주",
             "충청북": "충북", "충청남": "충남", "전라북": "전북", "전라남": "전남",
             "경상북": "경북", "경상남": "경남", "충청": "충북"}
    s = alias.get(s, s)
    return s if s in CENTROIDS else (OFFICE_DISTRICTS.get(str(v).strip()) if str(v).strip() in OFFICE_DISTRICTS else (s if s in CENTROIDS else None))

def load_xlsx(path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

def header_index(header_row, name):
    for i, h in enumerate(header_row):
        if h and name in str(h).replace("\n", " "):
            return i
    return None

def to_int(v):
    try:
        return int(float(v))
    except Exception:
        return 0

def agg_sheet_xlsx(ws, header_at, name_col, region_col, monitor_col, household_col=None, population_col=None, price_col=None):
    rows = list(ws.iter_rows(min_row=header_at, values_only=True))
    hdr = rows[0]
    ni = header_index(hdr, name_col); ri = header_index(hdr, region_col)
    mi = header_index(hdr, monitor_col)
    hi = header_index(hdr, household_col) if household_col else None
    pi = header_index(hdr, population_col) if population_col else None
    pri = header_index(hdr, price_col) if price_col else None
    complexes = 0; monitors = 0; households = 0; population = 0; prices = set()
    regions = {}
    for r in rows[1:]:
        if ni is None or ni >= len(r) or not r[ni]:
            continue
        complexes += 1
        m = to_int(r[mi]) if mi is not None and mi < len(r) else 0
        monitors += m
        if hi is not None and hi < len(r):
            households += to_int(r[hi])
        if pi is not None and pi < len(r):
            population += to_int(r[pi])
        if pri is not None and pri < len(r) and r[pri]:
            prices.add(to_int(r[pri]))
        sido = norm_sido(r[ri]) if ri is not None and ri < len(r) else None
        if sido:
            g = regions.setdefault(sido, {"complexes": 0, "monitors": 0})
            g["complexes"] += 1; g["monitors"] += m
    return dict(complexes=complexes, monitors=monitors, households=households,
                population=population, prices=sorted(p for p in prices if p), regions=regions)

def merge(a, b):
    out = dict(complexes=a["complexes"] + b["complexes"], monitors=a["monitors"] + b["monitors"],
               households=a["households"] + b["households"], population=a["population"] + b["population"],
               prices=sorted(set(a["prices"]) | set(b["prices"])), regions={})
    for src in (a["regions"], b["regions"]):
        for k, v in src.items():
            g = out["regions"].setdefault(k, {"complexes": 0, "monitors": 0})
            g["complexes"] += v["complexes"]; g["monitors"] += v["monitors"]
    return out

def build_townboard():
    wb = load_xlsx(os.path.join(SRC, "T사_아파트 엘리베이터.xlsx"))
    a = agg_sheet_xlsx(wb["타운보드S(전국 50,000대)"], 6, "아파트명", "지역1", "가동수량", "세대수")
    b = agg_sheet_xlsx(wb["타운보드L(전국 10,000대)"], 6, "아파트명", "지역1", "가동수량", "세대수")
    wb.close()
    return merge(a, b)

def build_fmk():
    wb = load_xlsx(os.path.join(SRC, "F사_아파트 엘리베이터.xlsx"))
    a = agg_sheet_xlsx(wb["FMK"], 4, "도시", "도시", "판매수량", "총 세대수", "총 인구수", "대당단가")
    wb.close()
    return a

def build_gsa():
    wb = load_xlsx(os.path.join(SRC, "G사_아파트 엘리베이터.xlsx"))
    a = agg_sheet_xlsx(wb["설치 리스트"], 5, "아파트명", "지역1", "합계", "세대수")
    wb.close()
    return a

def build_primeliving():
    """프라임리빙(spaceAdd) — 패키지 전용. 68빌딩·521기·세대/유동인구 집계."""
    wb = load_xlsx(os.path.join(SRC, "[spaceAdd] 2월_프라임리빙_판매안 리스트(대외용)_260107.xlsx"))
    ws = wb["판매안_프라임리빙_26년 2월"]
    data = list(ws.iter_rows(min_row=21, values_only=True))
    hdr = data[0]
    def gi(name):
        return header_index(hdr, name)
    ni, ai = gi("빌딩명"), gi("주소")
    ei, ii, hi = gi("E/V 외부"), gi("E/V 내부"), gi("세대수")
    complexes = monitors = households = 0
    regions = {}
    for r in data[1:]:
        if ni is None or ni >= len(r) or not r[ni]:
            continue
        addr = r[ai] if ai is not None and ai < len(r) else ""
        if not addr:
            continue
        complexes += 1
        m = to_int(r[ei] if ei is not None and ei < len(r) else 0) + to_int(r[ii] if ii is not None and ii < len(r) else 0)
        monitors += m
        households += to_int(r[hi] if hi is not None and hi < len(r) else 0)
        sido = norm_sido(str(addr).split()[0]) or "서울"
        g = regions.setdefault(sido, {"complexes": 0, "monitors": 0})
        g["complexes"] += 1
        g["monitors"] += m
    wb.close()
    return dict(complexes=complexes, monitors=monitors, households=households,
                population=0, prices=[], regions=regions)

def build_officebiz():
    wb = xlrd.open_workbook(os.path.join(SRC, "H_오피스 엘리베이터.xls"))
    ws = wb.sheet_by_name("오피스보드")
    hdr = [str(ws.cell_value(4, j)).replace("\n", " ") for j in range(ws.ncols)]
    def idx(name, start=0):
        for j in range(start, len(hdr)):
            if name in hdr[j]:
                return j
        return None
    ni = idx("건물명"); ri = idx("권역"); mi = idx("모니터 합계"); ti = idx("입주업체"); pi = idx("15초")
    complexes = monitors = tenants = 0; prices = set(); regions = {}
    for i in range(5, ws.nrows):
        name = ws.cell_value(i, ni) if ni is not None else None
        if not name:
            continue
        complexes += 1
        m = to_int(ws.cell_value(i, mi)) if mi is not None else 0
        monitors += m
        if ti is not None:
            tenants += to_int(ws.cell_value(i, ti))
        if pi is not None and ws.cell_value(i, pi):
            prices.add(to_int(ws.cell_value(i, pi)))
        sido = norm_sido(ws.cell_value(i, ri)) if ri is not None else "서울"
        sido = sido or "서울"
        g = regions.setdefault(sido, {"complexes": 0, "monitors": 0})
        g["complexes"] += 1; g["monitors"] += m
    return dict(complexes=complexes, monitors=monitors, households=0, population=tenants,
                prices=sorted(p for p in prices if p), regions=regions)

def build_asa():
    wb = load_xlsx(os.path.join(SRC, "A사_오피스 엘리베이터.xlsx"))
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    bases = [1, 7, 13, 19]
    complexes = monitors = 0; regions = {}
    carry = {b: None for b in bases}
    for r in rows:
        for b in bases:
            if b + 4 >= len(r):
                continue
            no, region, name, ext, intr = r[b], r[b + 1], r[b + 2], r[b + 3], r[b + 4]
            if region:
                carry[b] = region
            if name and isinstance(name, str) and name.strip() and name.strip() not in ("빌딩명", "No."):
                complexes += 1
                m = to_int(ext) + to_int(intr)
                monitors += m
                sido = norm_sido(carry[b]) or "서울"
                g = regions.setdefault(sido, {"complexes": 0, "monitors": 0})
                g["complexes"] += 1; g["monitors"] += m
    wb.close()
    return dict(complexes=complexes, monitors=monitors, households=0, population=0, prices=[], regions=regions)

NETWORKS = [
    # vendor = 매체사 코드(고정), brand = 상품명, saleUnit = 판매 단위(site=개별 단지/건물, package=패키지 전용)
    dict(id="townboard", vendor="T사", brand="타운보드", type="apartment", saleUnit="site",
         spec="25형 · 55형 게시판", placement="엘리베이터 내부 게시판",
         target="아파트 주거 가구 · 출입 동선 반복 노출",
         highlights=["전국 최대 커버리지", "가구 도달 규모 1위"], build=build_townboard),
    dict(id="fmk", vendor="F사", brand="포커스", type="apartment", saleUnit="site",
         spec="21.5형 · 25형", placement="엘리베이터 내부 · 대기공간",
         target="수도권 아파트·주상복합 주거민",
         highlights=["대당 단가 공개", "인구 도달 규모 최다"], build=build_fmk),
    dict(id="gsa", vendor="G사", brand="미디어믿", type="apartment", saleUnit="site",
         spec="게시판형", placement="엘리베이터 내부 · 대기공간",
         target="수도권 집중 아파트 주거민",
         highlights=["서울·경기 집중"], build=build_gsa),
    dict(id="officebiz", vendor="H사", brand="오피스비즈TV", type="office", saleUnit="site",
         spec="A타입 25형 · B타입 21.5형", placement="엘리베이터 내부 · 대기공간 · 디지털게시판",
         target="오피스 상주 직장인 · 방문 고객",
         highlights=["15초 단가 공개", "프라임 오피스 밀집"], build=build_officebiz),
    # spaceAdd 프라임리빙 — 주거(주상복합·레지던스·오피스텔), 패키지 전용
    dict(id="primeliving", vendor="A사", brand="프라임리빙", type="apartment", saleUnit="package",
         spec="E/V 외부 · 내부", placement="엘리베이터 외부 · 내부",
         target="도심 주상복합·레지던스 거주 직장인 가구",
         highlights=["CBD·GBD 핵심 권역", "패키지 전용"], build=build_primeliving),
    # spaceAdd 프라임오피스 — 패키지 전용(개별 가격 없음)
    dict(id="asa", vendor="A사", brand="프라임오피스", type="office", saleUnit="package",
         spec="내부 · 외부 모니터", placement="엘리베이터 내부 · 외부",
         target="CBD·GBD·YBD 오피스 직장인",
         highlights=["핵심 업무권역(CBD/GBD/YBD) 커버"], build=build_asa),
]

def main():
    networks = []
    for cfg in NETWORKS:
        data = cfg["build"]()
        regions = [dict(name=k, complexes=v["complexes"], monitors=v["monitors"],
                        centroid=CENTROIDS.get(k))
                   for k, v in sorted(data["regions"].items(), key=lambda kv: -kv[1]["monitors"])
                   if CENTROIDS.get(k)]
        total_m = data["monitors"] or 1
        top = regions[0]["name"] if regions else None
        top_share = round(regions[0]["monitors"] / total_m * 100) if regions else 0
        networks.append(dict(
            id=cfg["id"], vendor=cfg["vendor"], brand=cfg["brand"], type=cfg["type"],
            saleUnit=cfg.get("saleUnit", "site"),
            spec=cfg["spec"], placement=cfg["placement"], target=cfg["target"],
            highlights=cfg["highlights"],
            complexes=data["complexes"], monitors=data["monitors"],
            households=data["households"], population=data["population"],
            prices=data["prices"], topRegion=top, topRegionShare=top_share,
            regions=regions,
        ))
        print(f"[{cfg['vendor']} {cfg['brand']}] 단지/빌딩={data['complexes']:,} 모니터={data['monitors']:,} "
              f"세대={data['households']:,} 인구/입주={data['population']:,} 단가={data['prices']} top={top}({top_share}%)")
    out = dict(
        note="엘리베이터 광고 네트워크 집계 데이터. 개별 매체 핀이 아닌 네트워크·지역 단위 집계.",
        centroids=CENTROIDS,
        networks=networks,
    )
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("WROTE", os.path.normpath(OUT))

if __name__ == "__main__":
    main()
