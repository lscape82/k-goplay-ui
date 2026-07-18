# -*- coding: utf-8 -*-
"""엘리베이터 광고 상품별 대표 건물 주소를 지역 분산 샘플링 후 네이버 지오코딩 → data/elevator-sites.json.
전체 13만+ 건은 비현실적이므로 상품별 상한(CAP)만큼 시도별 라운드로빈으로 뽑아 실제 좌표를 얻는다.
CAP를 늘리면 커버리지를 확장할 수 있다."""
import json, os, re, time, urllib.parse, urllib.request, urllib.error
import openpyxl, xlrd

SRC = os.environ.get("ELEV_SRC", r"E:\2. 광고플레이(주)\★★★플랫폼개발2_2026_업데이트_언니,부장님\★★★플랫폼UI리모델링\2026 UI활용데이터\엘리베이터")
ROOT = os.path.join(os.path.dirname(__file__), "..")
OUT = os.path.join(ROOT, "data", "elevator-sites.json")
CAP = int(os.environ.get("ELEV_CAP", "45"))

def load_env():
    creds = {}
    p = os.path.join(ROOT, ".env")
    if os.path.exists(p):
        for line in open(p, encoding="utf-8"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                creds[k.strip()] = v.strip()
    return creds

ENV = load_env()
CID = ENV.get("NAVER_MAPS_CLIENT_ID")
CSECRET = ENV.get("NAVER_MAPS_CLIENT_SECRET")

METRO = ["서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종"]

def norm_sido(v):
    if not v:
        return "기타"
    s = str(v).strip()
    if s.startswith("전남광주통합"):  # T사 원본의 손상 라벨 = 광주광역시
        return "광주"
    for m in METRO:  # 서울시/서울특별시/세종특별자치시 등 광역시·특별시 계열
        if s.startswith(m):
            return m
    for t in ["특별자치도", "특별자치시", "특별시", "광역시"]:
        s = s.replace(t, "")
    if s.endswith("도"):
        s = s[:-1]
    alias = {"강원특별": "강원", "전북특별": "전북", "제주특별": "제주",
             "충청북": "충북", "충청남": "충남", "전라북": "전북", "전라남": "전남",
             "경상북": "경북", "경상남": "경남"}
    return alias.get(s, s)

def derive_region(address):
    """지오코딩된 주소에서 sido·district(구/군/시)를 일관되게 도출.
    예) '서울특별시 마포구 성암로 41' -> ('서울','마포구'), '경기도 수원시 영통구 …' -> ('경기','수원시')"""
    toks = str(address or "").split()
    sido = norm_sido(toks[0]) if toks else "기타"
    district = ""
    for t in toks[1:]:
        if t.endswith(("구", "군", "시")):
            district = t
            break
    if not district:
        district = sido  # 세종 등 하위 구 없는 경우
    return sido, district

def hidx(hdr, name):
    for i, h in enumerate(hdr):
        if h and name in str(h).replace("\n", " "):
            return i
    return None

def to_int(v):
    try:
        return int(float(v))
    except Exception:
        return 0

def cell(r, i):
    return r[i] if i is not None and i < len(r) else None

def rows_xlsx(path, sheet, header_at, name_c, region_c, addr_c, monitor_c,
              household_c=None, year_c=None, unitprice_c=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    data = list(ws.iter_rows(min_row=header_at, values_only=True))
    hdr = data[0]
    ni, ri, ai, mi = hidx(hdr, name_c), hidx(hdr, region_c), hidx(hdr, addr_c), hidx(hdr, monitor_c)
    hi = hidx(hdr, household_c) if household_c else None
    yi = hidx(hdr, year_c) if year_c else None
    pi = hidx(hdr, unitprice_c) if unitprice_c else None
    out = []
    for r in data[1:]:
        if ni is None or ni >= len(r) or not r[ni]:
            continue
        addr = cell(r, ai)
        if not addr:
            continue
        monitors = to_int(cell(r, mi))
        unitprice = to_int(cell(r, pi))
        out.append(dict(name=str(r[ni]).strip(), sido=norm_sido(cell(r, ri)),
                        address=str(addr).strip(), monitors=monitors,
                        households=to_int(cell(r, hi)), year=to_int(cell(r, yi)),
                        price=unitprice * monitors if unitprice else 0))
    wb.close()
    return out

def rows_townboard(path):
    """타운보드 로컬상품 — 요청 필드 전량 추출(구분·아파트명·입주년도·주소·평형·세대수·모니터수·개별단가·월비용·모니터크기)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in ["타운보드S(전국 50,000대)", "타운보드L(전국 10,000대)"]:
        if sheet not in wb.sheetnames:
            continue
        data = list(wb[sheet].iter_rows(min_row=6, values_only=True))
        hdr = data[0]
        c = {k: hidx(hdr, k) for k in ["구분", "아파트명", "입주년도", "지역1", "주소", "평형", "세대수", "가동수량", "개별 단가", "모니터크기"]}
        mc = hidx(hdr, "월비용")
        if mc is None:
            mc = hidx(hdr, "총 단가")
        for r in data[1:]:
            name, addr = cell(r, c["아파트명"]), cell(r, c["주소"])
            if not name or not addr:
                continue
            monitors = to_int(cell(r, c["가동수량"]))
            unit = to_int(cell(r, c["개별 단가"]))
            monthly = to_int(cell(r, mc)) or unit * monitors
            out.append(dict(
                gubun=str(cell(r, c["구분"]) or "아파트").strip(),
                name=str(name).strip(),
                year=to_int(cell(r, c["입주년도"])),
                sido=norm_sido(cell(r, c["지역1"])),
                address=str(addr).strip(),
                pyeong=str(cell(r, c["평형"]) or "").strip(),
                households=to_int(cell(r, c["세대수"])),
                monitors=monitors,
                unitPrice=unit,
                monthlyCost=monthly,
                monitorSize=str(cell(r, c["모니터크기"]) or "").strip(),
            ))
    wb.close()
    return out

def rows_fmk(path):
    """포커스(F사) — FMK 시트만 사용. 가격은 '4주 금액' 그대로(월 환산 금지)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["FMK"]
    data = list(ws.iter_rows(min_row=4, values_only=True))
    hdr = data[0]
    c = {k: hidx(hdr, k) for k in ["단지명", "도시", "주소(도로명)", "건물유형", "준공연도",
                                   "기준평형", "총 세대수", "판매수량", "대당단가", "4주 금액"]}
    out = []
    for r in data[1:]:
        name, addr = cell(r, c["단지명"]), cell(r, c["주소(도로명)"])
        if not name or not addr:
            continue
        monitors = to_int(cell(r, c["판매수량"]))
        if monitors <= 0:      # 판매수량 0 = 판매 구좌 없음 → 제외
            continue
        out.append(dict(
            gubun=str(cell(r, c["건물유형"]) or "아파트").strip(),
            name=str(name).strip(), year=to_int(cell(r, c["준공연도"])),
            sido=norm_sido(cell(r, c["도시"])), address=str(addr).strip(),
            pyeong=str(cell(r, c["기준평형"]) or "").strip(),
            households=to_int(cell(r, c["총 세대수"])), monitors=monitors,
            unitPrice=to_int(cell(r, c["대당단가"])),
            monthlyCost=to_int(cell(r, c["4주 금액"])),  # 4주 금액 그대로
        ))
    wb.close()
    return out

def rows_gsa(path):
    """미디어믿(G사) — 15초/30초 2축 단가."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["설치 리스트"]
    data = list(ws.iter_rows(min_row=5, values_only=True))
    hdr = data[0]
    c = {k: hidx(hdr, k) for k in ["구분", "아파트명", "지역1", "주 소", "합계", "준공년도",
                                   "최소평형", "최대평형", "세대수", "15초 단가", "월 광고료(15초)",
                                   "30초 단가", "월 광고료(30초)"]}
    out = []
    for r in data[1:]:
        name, addr = cell(r, c["아파트명"]), cell(r, c["주 소"])
        if not name or not addr:
            continue
        lo, hi = to_int(cell(r, c["최소평형"])), to_int(cell(r, c["최대평형"]))
        pyeong = f"{lo}~{hi}" if lo and hi and lo != hi else (str(lo) if lo else "")
        out.append(dict(
            gubun=str(cell(r, c["구분"]) or "아파트").strip(),
            name=str(name).strip(), year=to_int(cell(r, c["준공년도"])),
            sido=norm_sido(cell(r, c["지역1"])), address=str(addr).strip(),
            pyeong=pyeong, households=to_int(cell(r, c["세대수"])),
            monitors=to_int(cell(r, c["합계"])),
            unitPrice=to_int(cell(r, c["15초 단가"])),
            monthlyCost=to_int(cell(r, c["월 광고료(15초)"])),
            unitPrice30=to_int(cell(r, c["30초 단가"])),
            monthlyCost30=to_int(cell(r, c["월 광고료(30초)"])),
        ))
    wb.close()
    return out

def rows_primeliving(path):
    """프라임리빙(spaceAdd) — 패키지 전용(가격 열 없음). 주거(주상복합·레지던스·오피스텔)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["판매안_프라임리빙_26년 2월"]
    data = list(ws.iter_rows(min_row=21, values_only=True))
    hdr = data[0]
    c = {k: hidx(hdr, k) for k in ["권역", "빌딩명", "주소", "층수", "E/V 외부", "E/V 내부", "세대수", "유동인구"]}
    out = []
    for r in data[1:]:
        name, addr = cell(r, c["빌딩명"]), cell(r, c["주소"])
        if not name or not addr:
            continue
        monitors = to_int(cell(r, c["E/V 외부"])) + to_int(cell(r, c["E/V 내부"]))
        addr = str(addr).strip()
        out.append(dict(
            gubun="주상복합", name=str(name).strip(), year=0,
            sido=norm_sido(addr.split()[0] if addr else None) or "서울",
            address=addr, pyeong="", households=to_int(cell(r, c["세대수"])),
            monitors=monitors, unitPrice=0, monthlyCost=0,
            district=str(cell(r, c["권역"]) or "").strip(),
            floors=str(cell(r, c["층수"]) or "").strip(),
            traffic=to_int(cell(r, c["유동인구"])),
        ))
    wb.close()
    return out

def rows_primeoffice(path):
    """프라임오피스(asa) 핀 소스 — F사 파일의 '프라임오피스' 시트(주소 보강용, 606빌딩).
    ※ 네트워크 집계(747·4,847)는 A사 정본에서 별도로 잡히고, 여기선 주소 있는 606건만 지오코딩."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["프라임오피스"]
    data = list(ws.iter_rows(min_row=21, values_only=True))
    hdr = data[0]
    c = {k: hidx(hdr, k) for k in ["권역", "빌딩명", "주소", "층수", "E/V 외부", "E/V 내부",
                                   "주요입주사", "유동인구", "단가"]}
    out = []
    for r in data[1:]:
        name, addr = cell(r, c["빌딩명"]), cell(r, c["주소"])
        if not name or not addr:
            continue
        addr = str(addr).strip()
        monitors = to_int(cell(r, c["E/V 외부"])) + to_int(cell(r, c["E/V 내부"]))
        out.append(dict(
            gubun="오피스", name=str(name).strip(), year=0, pyeong="",
            sido=norm_sido(addr.split()[0] if addr else None) or "서울", address=addr,
            households=0, monitors=monitors, unitPrice=to_int(cell(r, c["단가"])), monthlyCost=0,
            district=str(cell(r, c["권역"]) or "").strip(),
            floors=str(cell(r, c["층수"]) or "").strip(),
            traffic=to_int(cell(r, c["유동인구"])),
            keyTenants=str(cell(r, c["주요입주사"]) or "").strip(),  # 텍스트(입주사 수 아님)
        ))
    wb.close()
    return out

def rows_asa(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    bases = [1, 7, 13, 19]
    carry = {b: None for b in bases}
    out = []
    for r in ws.iter_rows(values_only=True):
        for b in bases:
            if b + 4 >= len(r):
                continue
            region, name, ext, intr = r[b + 1], r[b + 2], r[b + 3], r[b + 4]
            if region:
                carry[b] = region
            if name and isinstance(name, str) and name.strip() and name.strip() not in ("빌딩명", "No."):
                nm = name.strip()
                q = re.sub(r"\s*[A-Da-d]?\d+동$", "", nm).strip()  # '1동','B동' 꼬리 제거
                out.append(dict(name=nm, sido="서울",
                                address=q, monitors=to_int(ext) + to_int(intr),
                                district=str(carry[b] or "")))
    wb.close()
    return out

def rows_h(path):
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name("오피스보드")
    hdr = [str(ws.cell_value(4, j)).replace("\n", " ") for j in range(ws.ncols)]
    def idx(name):
        for j, h in enumerate(hdr):
            if name in h:
                return j
        return None
    ni, ai, mi = idx("건물명"), idx("상세주소"), idx("모니터 합계")
    tenant_i, price_i = idx("입주업체"), idx("1개월 금액")
    unit_i, size_i = idx("15초 단가"), idx("모니터 형태")
    out = []
    for i in range(5, ws.nrows):
        name = ws.cell_value(i, ni) if ni is not None else None
        addr = ws.cell_value(i, ai) if ai is not None else None
        if not name or not addr:
            continue
        out.append(dict(gubun="오피스", name=str(name).strip(), sido="서울",
                        address=str(addr).strip(), year=0, pyeong="",
                        monitors=to_int(ws.cell_value(i, mi)) if mi is not None else 0,
                        households=0,
                        unitPrice=to_int(ws.cell_value(i, unit_i)) if unit_i is not None else 0,
                        monthlyCost=to_int(ws.cell_value(i, price_i)) if price_i is not None else 0,
                        monitorSize=str(ws.cell_value(i, size_i)).strip() if size_i is not None else "",
                        tenants=to_int(ws.cell_value(i, tenant_i)) if tenant_i is not None else 0))
    return out

def sample_spread(rows, cap):
    """시도별 라운드로빈 + 시도 내 모니터 많은 순 우선 → 지역 분산 샘플."""
    by = {}
    for r in rows:
        by.setdefault(r["sido"], []).append(r)
    for k in by:
        by[k].sort(key=lambda x: -x["monitors"])
    picked, i = [], 0
    order = sorted(by.keys(), key=lambda k: -len(by[k]))
    while len(picked) < cap and any(i < len(by[k]) for k in order):
        for k in order:
            if i < len(by[k]):
                picked.append(by[k][i])
                if len(picked) >= cap:
                    break
        i += 1
    return picked

def geocode(query):
    url = "https://naveropenapi.apigw.ntruss.com/map-geocode/v2/geocode?query=" + urllib.parse.quote(query)
    req = urllib.request.Request(url, headers={"x-ncp-apigw-api-key-id": CID, "x-ncp-apigw-api-key": CSECRET})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        addrs = data.get("addresses") or []
        if not addrs:
            return None
        a = addrs[0]
        return dict(lat=float(a["y"]), lng=float(a["x"]),
                    roadAddress=a.get("roadAddress", ""), jibunAddress=a.get("jibunAddress", ""))
    except Exception as exc:
        print("  geocode err:", str(exc)[:80])
        return None

PRODUCTS = [
    # 정본: T사_아파트 엘리베이터.xlsx (build-elevator-data.py와 동일 파일)
    dict(id="townboard", cap=None,  # 전량 지오코딩
         fn=lambda: rows_townboard(os.path.join(SRC, "T사_아파트 엘리베이터.xlsx"))),
    dict(id="fmk", cap=None,
         fn=lambda: rows_fmk(os.path.join(SRC, "F사_아파트 엘리베이터.xlsx"))),
    dict(id="gsa", cap=None,
         fn=lambda: rows_gsa(os.path.join(SRC, "G사_아파트 엘리베이터.xlsx"))),
    dict(id="officebiz", cap=None,
         fn=lambda: rows_h(os.path.join(SRC, "H_오피스 엘리베이터.xls"))),
    dict(id="primeliving", cap=None,
         fn=lambda: rows_primeliving(os.path.join(SRC, "[spaceAdd] 2월_프라임리빙_판매안 리스트(대외용)_260107.xlsx"))),
    # 프라임오피스 핀 — A사엔 주소 없음 → F사 '프라임오피스' 시트로 주소 보강(606/747, 나머지 141 주소 부재)
    dict(id="asa", cap=None,
         fn=lambda: rows_primeoffice(os.path.join(SRC, "F사_아파트 엘리베이터.xlsx"))),
]

ONLY = os.environ.get("ELEV_ONLY", "").strip()  # 특정 상품만 재생성(쉼표구분)

def main():
    if not CID or not CSECRET:
        raise SystemExit("NAVER_MAPS_CLIENT_ID/SECRET 미설정(.env 확인)")
    # 기존 데이터 유지 후, 파일이 있는 상품만 갱신(다른 상품 원본은 삭제됨)
    result = {}
    if os.path.exists(OUT):
        result = json.load(open(OUT, encoding="utf-8"))
    only = set(x for x in ONLY.split(",") if x)
    for p in PRODUCTS:
        if only and p["id"] not in only:
            continue
        try:
            rows = p["fn"]()
        except FileNotFoundError:
            print(f"[{p['id']}] 원본 파일 없음 → 기존 {len(result.get(p['id'], []))}개 유지")
            continue
        cap = p.get("cap", CAP)
        sample = rows if cap is None else sample_spread(rows, cap)
        print(f"[{p['id']}] 전체 {len(rows):,} → {'전량' if cap is None else '샘플'} {len(sample)} 지오코딩...")
        sites = []
        for idx, s in enumerate(sample):
            g = geocode(s["address"])
            time.sleep(0.08)
            if not g:
                continue
            site = dict(s)
            site["id"] = f"{p['id']}-{idx}"
            site["address"] = g["roadAddress"] or s["address"]
            site["lat"], site["lng"] = g["lat"], g["lng"]
            # sido·district를 주소에서 재도출(하드코딩 sido·비표준값 교정 + 구 단위 클러스터용)
            site["sido"], site["district"] = derive_region(site["address"])
            sites.append(site)
            if (idx + 1) % 300 == 0:
                print(f"    ...{idx + 1}/{len(sample)} (좌표 {len(sites)})")
        result[p["id"]] = sites
        print(f"    → 좌표 확보 {len(sites)}개")
        # 중간 저장(장시간 작업 안전)
        with open(OUT, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
    total = sum(len(v) for v in result.values())
    print(f"WROTE {os.path.normpath(OUT)}  (총 {total} 사이트)")

if __name__ == "__main__":
    main()
