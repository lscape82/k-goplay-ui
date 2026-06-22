from __future__ import annotations

import json
import csv
import re
from pathlib import Path
from typing import Any

import openpyxl


SOURCE_KEY = "\uc11c\uc6b8\uc2dc\ubc84\uc2a4\uc815\ub958\uc18c\uc704\uce58\uc815\ubcf4"
COST_KEY = "\ub4f1\uae09\ubcc4 \ube44\uc6a9"
OUTPUT_PATH = Path("data/bus_stops.json")
INSTALLATION_COST = 80_000
RIDERSHIP_KEY = "BUS_STATION_BOARDING_MONTH"


def find_workbook(cost: bool) -> Path:
    desktop = Path.home() / "Desktop"
    matches = [
        path
        for path in desktop.glob("*.xlsx")
        if SOURCE_KEY in path.name
        and (COST_KEY in path.name) == cost
        and not path.name.startswith("~$")
    ]
    if not matches:
        label = "cost" if cost else "location"
        raise FileNotFoundError(f"Bus stop {label} workbook was not found on Desktop.")
    return sorted(matches, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def find_ridership_csv() -> Path | None:
    desktop = Path.home() / "Desktop"
    matches = [
        path
        for path in desktop.glob("*.csv")
        if RIDERSHIP_KEY in path.name
        and not path.name.startswith("~$")
    ]
    return sorted(matches, key=lambda path: path.stat().st_mtime, reverse=True)[0] if matches else None


def normalize_ars(value: object) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) > 5:
        digits = digits[:-1]
    if len(digits) > 5:
        digits = digits[:5]
    return digits.zfill(5) if digits else ""


def parse_price(value: object) -> int | None:
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"\D", "", str(value or ""))
    return int(digits) if digits else None


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def classify_product(light_values: list[str], type_values: list[str]) -> str:
    text = " ".join(light_values + type_values)
    if any(keyword in text.upper() for keyword in ("DID", "LED")) or "\ub514\uc9c0\ud138" in text or "\uc601\uc0c1" in text:
        return "digital"
    if "\ube44\uc870\uba85" in text:
        return "nonlight"
    if "\uc870\uba85" in text:
        return "light"
    return "static"


def display_label(kind: str) -> str:
    return {
        "digital": "\ub514\uc9c0\ud138 \uc601\uc0c1\ud615",
        "light": "\uc870\uba85 \uace0\uc815\ud615",
        "nonlight": "\ube44\uc870\uba85 \uace0\uc815\ud615",
        "static": "\uace0\uc815\ud615",
    }.get(kind, "\uace0\uc815\ud615")


def unique(values: list[str]) -> list[str]:
    return sorted({value for value in values if value})


def merge_cost_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    prices = [price for price in (parse_price(row.get("monthlyCost")) for row in rows) if price is not None]
    grades = unique([clean_text(row.get("grade")) for row in rows])
    light_values = unique([clean_text(row.get("light")) for row in rows])
    type_values = unique([clean_text(row.get("type")) for row in rows])
    operators = unique([clean_text(row.get("operator")) for row in rows])
    faces = unique([clean_text(row.get("face")) for row in rows])
    districts = unique([clean_text(row.get("district")) for row in rows])
    dongs = unique([clean_text(row.get("dong")) for row in rows])
    station_names = unique([clean_text(row.get("stationName")) for row in rows])
    addresses = unique([clean_text(row.get("address")) for row in rows])
    inner_sizes = unique([clean_text(row.get("innerSize")) for row in rows])
    outer_sizes = unique([clean_text(row.get("outerSize")) for row in rows])
    kinds = classify_product(light_values, type_values)
    monthly_cost = min(prices) if prices else None
    return {
        "operator": ", ".join(operators),
        "displayKind": kinds,
        "displayLabel": display_label(kinds),
        "district": districts[0] if len(districts) == 1 else " / ".join(districts),
        "dong": dongs[0] if len(dongs) == 1 else " / ".join(dongs),
        "productType": type_values[0] if len(type_values) == 1 else " / ".join(type_values),
        "grade": grades[0] if len(grades) == 1 else " / ".join(grades),
        "stationName": station_names[0] if station_names else "",
        "faces": faces,
        "faceLabel": ", ".join(faces),
        "address": addresses[0] if len(addresses) == 1 else " / ".join(addresses),
        "light": " / ".join(light_values),
        "minMonthlyCost": monthly_cost,
        "maxMonthlyCost": max(prices) if prices else None,
        "monthlyCostLabel": format_won(monthly_cost) if monthly_cost else "",
        "installationCost": INSTALLATION_COST,
        "installationCostLabel": format_won(INSTALLATION_COST),
        "innerSize": inner_sizes[0] if len(inner_sizes) == 1 else " / ".join(inner_sizes),
        "outerSize": outer_sizes[0] if len(outer_sizes) == 1 else " / ".join(outer_sizes),
        "rows": rows,
    }


def format_won(value: int | None) -> str:
    if not value:
        return ""
    return f"{value:,}\uc6d0"


def format_people(value: int | None) -> str:
    if not value:
        return ""
    return f"{value:,}\uba85"


def read_ridership() -> dict[str, int]:
    source = find_ridership_csv()
    if not source:
        return {}
    totals: dict[str, int] = {}
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            with source.open(encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                for row in reader:
                    ars = normalize_ars(row.get("ARS_ID"))
                    if not ars:
                        continue
                    board = parse_price(row.get("\uc2b9\ucc28\ucd1d\uc2b9\uac1d\uc218")) or 0
                    alight = parse_price(row.get("\ud558\ucc28\ucd1d\uc2b9\uac1d\uc218")) or 0
                    totals[ars] = totals.get(ars, 0) + board + alight
            return totals
        except UnicodeDecodeError:
            continue
    return totals


def read_costs() -> dict[str, dict[str, Any]]:
    source = find_workbook(cost=True)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    configs = [
        {
            "sheet": "KT\uac00\ub85c\ubcc0",
            "start_row": 2,
            "operator": "KT",
            "columns": {
                "ars": 0,
                "district": 1,
                "dong": 2,
                "type": 3,
                "grade": 4,
                "stationName": 5,
                "face": 6,
                "address": 7,
                "light": 8,
                "monthlyCost": 10,
                "innerSize": 11,
                "outerSize": 12,
            },
        },
        {
            "sheet": "\uce74\uce74\uc624 \uac00\ub85c\ubcc0+\uc911\uc559",
            "start_row": 4,
            "operator": "\uce74\uce74\uc624\ubaa8\ube4c\ub9ac\ud2f0",
            "columns": {
                "ars": 1,
                "district": 2,
                "dong": 3,
                "type": 4,
                "grade": 5,
                "stationName": 6,
                "face": 7,
                "address": 8,
                "light": 9,
                "monthlyCost": 10,
                "outerSize": 11,
                "innerSize": 12,
            },
        },
    ]
    grouped: dict[str, list[dict[str, Any]]] = {}
    for config in configs:
        if config["sheet"] not in workbook.sheetnames:
            continue
        worksheet = workbook[config["sheet"]]
        columns = config["columns"]
        for row in worksheet.iter_rows(min_row=config["start_row"], values_only=True):
            ars = normalize_ars(row[columns["ars"]] if columns["ars"] < len(row) else "")
            if not ars:
                continue
            item: dict[str, Any] = {"operator": config["operator"], "ars": ars}
            for key, index in columns.items():
                if key == "ars":
                    continue
                item[key] = row[index] if index < len(row) else ""
            item["installationCost"] = INSTALLATION_COST
            grouped.setdefault(ars, []).append(item)
    return {ars: merge_cost_rows(rows) for ars, rows in grouped.items()}


def normalize_stop(row: dict[str, object]) -> dict[str, Any] | None:
    try:
        longitude = float(row["X\uc88c\ud45c"])
        latitude = float(row["Y\uc88c\ud45c"])
    except (TypeError, ValueError, KeyError):
        return None

    if not (37 <= latitude <= 38 and 126 <= longitude <= 128):
        return None

    node_id = clean_text(row.get("NODE_ID"))
    ars_id = normalize_ars(row.get("ARS_ID"))
    name = clean_text(row.get("\uc815\ub958\uc18c\uba85"))
    stop_type = clean_text(row.get("\uc815\ub958\uc18c\ud0c0\uc785"))
    if not node_id or not name:
        return None

    return {
        "id": node_id,
        "ars": ars_id,
        "name": name,
        "latitude": round(latitude, 7),
        "longitude": round(longitude, 7),
        "type": stop_type,
    }


def main() -> None:
    source = find_workbook(cost=False)
    costs = read_costs()
    ridership = read_ridership()
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]

    stops: list[dict[str, Any]] = []
    seen: set[str] = set()
    for values in rows:
        row = dict(zip(headers, values))
        stop = normalize_stop(row)
        if not stop:
            continue
        key = str(stop["id"])
        if key in seen:
            continue
        product = costs.get(str(stop["ars"]))
        if product:
            product = dict(product)
            monthly_riders = ridership.get(str(stop["ars"]))
            if monthly_riders:
                product["ridershipTotal"] = monthly_riders
                product["ridershipLabel"] = format_people(monthly_riders)
            stop["adProduct"] = product
        seen.add(key)
        stops.append(stop)

    stops.sort(key=lambda stop: (str(stop.get("adProduct", {}).get("district", "")), str(stop["name"]), str(stop["ars"])))
    OUTPUT_PATH.write_text(json.dumps(stops, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    matched = sum(1 for stop in stops if stop.get("adProduct"))
    print(f"Wrote {len(stops):,} bus stops ({matched:,} ad stops) to {OUTPUT_PATH} from {source.name}")


if __name__ == "__main__":
    main()
